import sys
import streamlit as st
import pandas as pd
import pprint
from datetime import datetime
from pathlib import Path
import base64
from openpyxl import load_workbook
import numpy as np
import openpyxl
import pandas as pd
import re
from datetime import datetime, time, timedelta



def normalize_product_name(name):
    if not name:
        return ""
    # Mise en majuscules et suppression des caractères spéciaux
    name = name.upper()
    
    # Supprime les espaces entre les lettres : "D A P" → "DAP"
    name = re.sub(r'\b([A-Z])\s+([A-Z])\b', r'\1\2', name)

    # Supprime tous les caractères non alphanumériques et remplace-les par un espace
    name = re.sub(r'[^A-Z0-9]', ' ', name)

    # Supprime les espaces multiples
    name = re.sub(r'\s+', ' ', name).strip()

    # Dictionnaire des règles de normalisation
    patterns = {
        'MAP 11 52 SPECIAL': [
            r'MAP\s*11\s*52',
            r'MAP\s*SPECIAL',
            r'MAP\s*11\s*52\s*SPECIAL',
            r'MAP\s*11\s*52\s*Special\s*Low\s*Cd',

        ],
        'NPK 14 18 18 6S 1B2O3': [
            r'NPK\s*14\s*18\s*18\s*6S\s*1B2O3\s*AFRIQUE',
            r'NPK\s*14\s*18\s*18\s*6S\s*1B2O3'],
         

        'DAP SPECIAL': [
            r'DAP\s*SPC',
            r'DAP\s*SPECIAL',
        ],
        'DAP EURO': [
            r'DAP\s*EURO',
            r'DAP\s*EU',
            r'DAP\s*Euro\s*Low\s*Cd'
            
        ],
        'DAP STANDARD': [
            r'DAP\s*STANDARD',
            r'DAP\s*STD',
        ],
        'TSP SPECIAL JORF': [
            r'TSP\s*JORF',
            r'TSP-JORF',
            r'TSP\s*SPECIAL\s*JORF',
            r'TSP\s*SPC\s*JARF',

        ],
        'NPS 3 30 9S': [
            r'NPS\s*3\s*30\s*9S\s*OFAS',
            r'NPS\s*3\s*30\s*9S',
        ],
        'UREE': [r'\bUREE\b', r'\bURÉE\b']
    }

    for standard, variants in patterns.items():
        for pattern in variants:
            if re.search(pattern, name):
                return standard

    return name# retourne le nom nettoyé mais non reconnu


# Configurer la mise en page
st.set_page_config(
    layout="wide",
    page_title="Suivi des chargements",
    page_icon="🌐",
)

# ✅ CSS : Barre verte + placement heure à droite
st.markdown("""
    <style>
        .top-bar {
            background-color: #00a65a; /* Vert */
            height: 4px;
            width: 100%;
            position: fixed;
            top: 0;
            left: 0;
            z-index: 100;
        }
            

        .custom-title {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding-top: 20px;
            padding-bottom: 10px;
            
        }

        .title-text {
            font-size: 1em;
            font-weight: bold;

        }

        .time-box {
            background-color: #85A98F;
            padding: 10px 20px;
            border-radius: 10px;
            font-size: 1em;
            color: #FFFFFF;
        }
    
        .block-container {
            padding-top: 30px !important;
        }
    </style>
    <div class="top-bar"></div>
""", unsafe_allow_html=True)

# ✅ Afficher Titre + Heure
now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

logo_path = "logo-white.png"  # Vérifie que le fichier est bien dans le même dossier

# Lecture du fichier et encodage en base64
logo_base64 = base64.b64encode(Path(logo_path).read_bytes()).decode()

st.markdown(f"""
    <div style="display: flex; align-items: center; justify-content: space-between; background-color: #328E6E; padding: 7px; border-radius: 8px;">
        <div style="flex: 1; display: flex; align-items: center;">
            <img src="data:image/png;base64,{logo_base64}" alt="logo" width="70"  style="margin-top : 10px; margin-right: 18px;"/>
        </div>
        <div style="flex: 2;color:#FFFFFF; text-align: center;margin-top : 10px; font-size: 24px; font-weight: bold;">
            🌐 Axe Chargement digital
        </div>
        <div style="flex: 1;color : White ; text-align: right; font-size: 18px;">
            🕒 {now}
        </div>
    </div>
""", unsafe_allow_html=True)

# Titre centré
# Charger les données depuis le fichier Excel

now = datetime.now()

# Si l'heure est avant 7h, utiliser la feuille du jour précédent
if now.time() < time(7, 0):
    effective_date = now - timedelta(days=1)
else:
    effective_date = now

# Formater la date au format "dd-mm-YYYY"
sheet_name = effective_date.strftime("%d-%m-%Y")

# Charger le fichier Excel et extraire la feuille souhaitée
excel_file = "SituationHFN.xlsx"
df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

# Extraire les données des quais
loading_data = {}

# Parcourir les colonnes de B à AD (index 1 à 30)
for col in range(1, 31):
    quai = df.iloc[6, col]  # Ligne 7 = index 6
    if pd.isna(quai):
        continue

    ship = df.iloc[1, col]             # Ligne 2 = index 1
    quantity_requested = df.iloc[4, col]  # Ligne 5 = index 4
    product_type = df.iloc[5, col]     # Ligne 6 = index 5
    origin = df.iloc[7, col]  
    total_charge = df.iloc[37, 0]  # Ligne 30, colonne B (0-indexed)

    # Cumul de chargement : lignes 13 à 36 (index 12 à 35)
    cumul_data = df.iloc[12:36, col].dropna()

    max_loaded = 0
    tonnage_last_hour = 0

    if not cumul_data.empty:
        max_index = cumul_data.idxmax()  # Ligne où le cumul est max
        max_loaded = cumul_data[max_index]

        # Vérifier que la colonne col+1 existe
        if col + 1 < df.shape[1]:
            tonnage_last_hour = df.iloc[max_index, col + 1]
    
    # Construction de la structure
    if quai not in loading_data:
        loading_data[quai] = {
            "ship": ship,
            "products": {}
        }

    loading_data[quai]["products"][product_type]= {
        "loaded": max_loaded,
        "target": quantity_requested,
        "last_hour": tonnage_last_hour,
        "Source" : origin,
        
    }
print(loading_data)
# Afficher les données dans le tableau de bord
page = st.sidebar.selectbox(
    "Navigation",
    ["Suivi de chargement", "Planification", "Stock","Navires en Rade","CTE"]
)

# Vue 1 : Suivi de chargement
# Vue 1 : Suivi de chargement
# Vue 1 : Suivi de chargement
if page == "Suivi de chargement":
    st.markdown(
    f"""
    <div style='display:flex; justify-content:space-between; align-items:center;'>
        <h3 style="font-size:24px; margin:0;">🚢 Suivi de chargement en temps réel</h3>
        <span style="font-size:20px;color:#FFFFFF; background-color:#F7374F; padding:6px 12px; border-radius:8px;">
            Total Chargé : <strong>{total_charge} t</strong>
        </span>
    </div>
    """,
    unsafe_allow_html=True
)



    # Diviser la page en deux colonnes : Quai 1 (gauche) et Quai 2 (droite)
    col1, col2 = st.columns(2)

    # Afficher les quais de gauche (1NORD, 1BIS, 1TER)
    with col1:
        st.markdown('<h4 style="margin-bottom:10px;">🧭 Quais - 1</h4>', unsafe_allow_html=True)
        for quai, info in loading_data.items():
            if quai.startswith("1"):  # Filtrer les quais de gauche
                # Vérifier si le quai a un navire
                if not info["ship"] or pd.isna(info["ship"]):  # Si pas de navire
                    st.markdown(f'<div style="font-size:18px; font-weight:bold;">Quai {quai} – 🚩 Quai Libre</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div style="font-size:18px; font-weight:bold;">Quai {quai} – Navire : {info["ship"]}</div>', unsafe_allow_html=True)
                    for product, stats in info["products"].items():
                        progress = stats["loaded"] / stats["target"] if stats["target"] > 0 else 0
                        percentage = progress * 100
                        source = stats["Source"]
                        if source == "JFC1":
                            color = "#FFEB3B"  # Jaune
                        elif source == "JFC2":
                           color = "#FFEB3B"  # Gris
                        elif source == "JLN":
                           color = "#FFEB3B"  # Jaune
                        else:
                           color = "#4CAF50"
                        st.markdown(f'<div style="font-size:14px;">🔸 Qualité 🌾 : {product} | Chargé ✅ : {stats["loaded"]} / {stats["target"]} t | Dernière heure 🕒 : {stats["last_hour"]} t </div>', unsafe_allow_html=True)
                        st.progress(progress)
                    
                        st.markdown(f'<div style="font-size:14px; text-align:center;">{percentage:.2f}% Chargé| Source de chargement actuel  🏗️ : {stats["Source"]}</div>', unsafe_allow_html=True)
                        
                st.markdown("<hr style='margin:10px 0;'>", unsafe_allow_html=True)

    # Afficher les quais de droite (2NORD, 2SUD, 2BIS, 2TER)
    with col2:
        st.markdown('<h4 style="margin-bottom:10px;">🧭 Quais - 2</h4>', unsafe_allow_html=True)
        for quai, info in loading_data.items():
            if quai.startswith("2"):  # Filtrer les quais de gauche
                # Vérifier si le quai a un navire
                if not info["ship"] or pd.isna(info["ship"]):  # Si pas de navire
                    st.markdown(f'<div style="font-size:18px; font-weight:bold;">Quai {quai} – 🚩 Quai Libre</div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div style="font-size:18px; font-weight:bold;">Quai {quai} – Navire : {info["ship"]}</div>', unsafe_allow_html=True)
                    for product, stats in info["products"].items():
                        progress = stats["loaded"] / stats["target"] if stats["target"] > 0 else 0
                        percentage = progress * 100
                        
                        st.markdown(f'<div style="font-size:14px;">🔸 Qualité 🌾 : {product} | Chargé ✅ : {stats["loaded"]} / {stats["target"]} t | Dernière heure 🕒 : {stats["last_hour"]} t </div>', unsafe_allow_html=True)
                        st.progress(progress)
                        st.markdown(f'<div style="font-size:14px; text-align:center;">{percentage:.2f}% Chargé| Source de chargement actuel  🏗️ : {stats["Source"]}</div>', unsafe_allow_html=True)
                st.markdown("<hr style='margin:10px 0;'>", unsafe_allow_html=True)

elif page == "Planification":

    if "stock_data" in st.session_state:
        # Use the stock data from session state
        combined_stock_data = st.session_state["stock_data"]
        print(combined_stock_data)
        st.success("✅ Données de stock importées depuis la page Stock.")
    else:
        st.warning("⚠️ Les données de stock ne sont pas disponibles. Veuillez d'abord les importer dans la page Stock.")
        st.stop()

    def filter_ships_by_axes(stock, navires, hangar_to_axes_usine, axe_usine_to_port, axe_Port_to_quai):
        filtered_results = []
        
        for navire in navires:
            navire_nom = navire["Navire"]
            quai = navire["Quai"]
            type_produit = navire["Type_Produit"]
            reste_a_charger = navire["Reste_A_Char"]
            
            sources_valides = []
            for source in stock:
                if source["Type de produit"] == type_produit and source["Quantité"] > 0:
                    hangar = source["Source"]
                    if hangar in hangar_to_axes_usine:
                        for axe_usine in hangar_to_axes_usine[hangar]:
                            if axe_usine in axe_usine_to_port:
                                for axe_port in axe_usine_to_port[axe_usine]:
                                    if axe_port in axe_Port_to_quai and quai in axe_Port_to_quai[axe_port]:
                                        sources_valides.append((source, axe_usine, axe_port))
            
            for source, axe_usine, axe_port in sources_valides:
                filtered_results.append({
                    "Navire": navire_nom,
                    "Quai": quai,
                    "Type_Produit": type_produit,
                    "rest_a_charger": reste_a_charger,
                    "Source": source["Source"],
                    "Quantite_Disponible": source["Quantité"],
                    "Axe_Usine": axe_usine,
                    "Axe_Port": axe_port
                })
        
        return filtered_results

    # Charger les données depuis Excel pour les navires et les stocks
    if 'navires_data' in st.session_state:
        navires_data = st.session_state['navires_data']
        print(navires_data)
        st.write("Using saved navires_data from session state")
    else:
        st.warning("No saved navires data found. Please upload the ship data in the 'Navires en Rade' page.")
        st.stop()
    navires = [ship for product in navires_data for ship in product["Ship Data"]]


    hangar_to_axes_usine = { 
        "18A": ["RB1", "RB2"], "18B": ["RB1", "RB2"], "18C": ["RB1", "RB2"],
        "HE01": ["RB1", "RB2"], "HE02": ["RB1", "RB2"], "HE03": ["RB1", "RB2"], "HE04": ["RB1", "RB2"],
        "HE05": ["RB3"], "HE06": ["RB3"],
        "H1BIS": ["RB1", "RB2", "RB3"], "HE2BIS": ["RB1", "RB2", "RB3"],
        "H3BIS": ["RB1", "RB2", "RB3"], "H4BIS": ["RB1", "RB2", "RB3"],
        "JFC1": ["TB1", "TB2"], "JFC3": ["TB1", "TB2"],
        "JFC4": ["TB1", "TB2", "TB3"], "JFC2": ["TB1", "TB2", "TB3"],
        "JFC5": ["TB1"], "107F": ["TB1"],
        "107D": ["TB1", "TB2", "TB3"], "107E": ["TB1", "TB2", "TB3"],"107F": ["TB1"],
    }

    axe_usine_to_port = {
        "RB1": ["G1", "G4", "GH3"], "RB2": ["G2", "G4", "GH3"], "RB3": ["GH13", "G3", "G4"],
        "TB1": ["G1", "G2", "G3", "G4", "GH3", "GH13", "GH4"],
        "TB2": ["G1", "G2", "G3", "G4", "GH3", "GH4", "GH13"],
        "TB3": ["GH4"]
    }

    axe_Port_to_quai = {
    "G1": ["1S", "1N"], "G2": ["1S", "1N"], "G3": ["1BIS", "1TER"], "G4": ["1BIS", "1TER"], 
    "GH4": ["2S", "2N", "2BIS", "2TER"], "GH13": ["2S", "2N", "2BIS", "2TER"],
    }

    # Filtrer les navires par les axes
    filtered_results = filter_ships_by_axes(combined_stock_data, navires, hangar_to_axes_usine, axe_usine_to_port, axe_Port_to_quai)

    scenarios = [
    {
            "navire": r["Navire"],  # Nom du navire
            "quai": r["Quai"],  # Quai
            "produit": r["Type_Produit"],  # Produit
            "quantite_restant": r["rest_a_charger"],  # Quantité restante à charger
            "quantite_stock": r["Quantite_Disponible"],  # Quantité disponible
            "source": r["Source"],  # Source
            "axe_usine": r["Axe_Usine"],  # Axe usine
            "axe_port": r["Axe_Port"],  # Axe port
        }
        for r in filtered_results
    ]

    def planifier_chargement(scenarios, debit_usine, max_heures=24, pause=1, controle_torrent_eau=1):
        planning = []
        navires_temps = {}  # Stocke le temps de début de chargement des navires
        axes_temps = {axe: [] for axe in debit_usine}  # Disponibilité des axes usine (liste des plages horaires utilisées)
        ports_temps = {}  # Disponibilité des axes port (liste des plages horaires utilisées)
        sources_temps = {}  # Disponibilité des sources
        stock_restant = {}  # Stock restant par source et produit
        navires_restant = {}  # Quantité restante à charger par navire et produit
        total_charge_journee = 0
        epuisement_signale = set()  # Ensemble pour suivre les épuisements déjà signalés

        # Initialiser les stocks restants et les besoins des navires
        for scenario in scenarios:
            key_stock = (scenario["source"], scenario["produit"])
            key_navire = (scenario["navire"], scenario["produit"])
            
            if key_stock not in stock_restant:
                stock_restant[key_stock] = scenario["quantite_stock"]
            
            if key_navire not in navires_restant:
                navires_restant[key_navire] = scenario["quantite_restant"]

        # Trier les scénarios pour maximiser le chargement
        scenarios.sort(key=lambda s: (-s["quantite_restant"], -s["quantite_stock"]))

        for scenario in scenarios:
            navire = scenario["navire"]
            produit = scenario["produit"]
            source = scenario["source"]
            axe_usine = scenario["axe_usine"]
            axe_port = scenario["axe_port"]
            key_stock = (source, produit)
            key_navire = (navire, produit)
            quai = scenario["quai"]

            # Vérification du stock disponible et de la demande du navire
            stock_dispo = stock_restant.get(key_stock, 0)
            demande_navire = navires_restant.get(key_navire, 0)
            if stock_dispo <= 0:
                # Enregistrer l'épuisement de stock une seule fois
                if key_stock not in epuisement_signale:
                    planning.append({
                        "navire": navire,
                        "produit": produit,
                        "source": source,
                        "quai": quai,
                        "message": "Épuisement de stock"
                    })
                    epuisement_signale.add(key_stock)
                continue

            if demande_navire <= 0:
                continue
            
            # Trouver le moment où l'axe et la source seront disponibles
            debut = max(navires_temps.get(navire, 0), sources_temps.get(source, 0))
            if sources_temps.get(source, 0) > navires_temps.get(navire, 0):
                debut = sources_temps[source] + pause

            heures_disponibles = max_heures - debut
            if heures_disponibles <= 0:
                continue

            # Vérifier si l'axe usine est disponible
            axe_usine_occupe = False
            for plage in axes_temps[axe_usine]:
                if not (debut >= plage[1] or (debut + (demande_navire / debit_usine[axe_usine])) <= plage[0]):
                    axe_usine_occupe = True
                    break

            if axe_usine_occupe:
                # Si l'axe usine est occupé, passer au prochain scénario
                continue

            # Vérifier si l'axe port est disponible
            if axe_port not in ports_temps:
                ports_temps[axe_port] = []  # Initialiser si non présent

            axe_port_occupe = False
            for plage in ports_temps[axe_port]:
                if not (debut >= plage[1] or (debut + (demande_navire / debit_usine[axe_usine])) <= plage[0]):
                    axe_port_occupe = True
                    break

            if axe_port_occupe:
                # Si l'axe port est occupé, passer au prochain scénario
                continue

            # Débit max possible sur l'axe sélectionné
            debit_max = debit_usine[axe_usine] * heures_disponibles
            quantite_chargee = min(demande_navire, debit_max, stock_dispo)
            if quantite_chargee > 0:
                fin_chargement = debut + (quantite_chargee / debit_usine[axe_usine])
                planning.append({
                    "navire": navire,
                    "quai": quai,
                    "produit": produit,
                    "debut": debut,
                    "fin": fin_chargement,
                    "quantite_chargee": quantite_chargee,
                    "axe_usine": axe_usine,
                    "axe_port": axe_port,
                    "source": source,
                    "message": f"Chargement du produit {produit}"
                })

                # Mise à jour des stocks et disponibilités
                stock_restant[key_stock] -= quantite_chargee
                navires_restant[key_navire] -= quantite_chargee
                navires_temps[navire] = fin_chargement + pause
                axes_temps[axe_usine].append((debut, fin_chargement))  # Ajouter la plage horaire utilisée pour l'axe usine
                ports_temps[axe_port].append((debut, fin_chargement))  # Ajouter la plage horaire utilisée pour l'axe port
                sources_temps[source] = fin_chargement + pause
                total_charge_journee += quantite_chargee
                

    

        return planning, total_charge_journee
    # Exemple de test
    debit_usine = {"RB1": 500, "RB2": 500, "RB3": 700, "TB1": 900, "TB2": 800, "TB3": 800}

    # Affichage des résultats
    planning, total_charge_journee = planifier_chargement(scenarios, debit_usine)

    for entry in planning:
        if "debut" in entry and "fin" in entry and "quantite_chargee" in entry:
            print(f"{entry['navire']} | {entry['produit']} | {entry['debut']:.2f}h | {entry['fin']:.2f}h | {entry['quantite_chargee']}t | {entry['axe_usine']} | {entry['axe_port']} | {entry['source']} | {entry.get('message', '')}")
        else:
            print(f"{entry['navire']} | {entry['produit']} | {entry.get('message', 'Information manquante')}")

    print(f"Total chargé dans la journée : {total_charge_journee} tonnes")
    st.subheader("📋 Planification de chargement par Quai")

    # Dictionnaire pour grouper par quai
    planning_par_quai = {}

    for entry in planning:
        quai = entry.get("quai", "Quai inconnu")
        if quai not in planning_par_quai:
            planning_par_quai[quai] = []
        
        # Préparer les infos pour chaque ligne
        if "debut" in entry and "fin" in entry and "quantite_chargee" in entry:
            planning_par_quai[quai].append({
                "Navire": entry["navire"],
                "Quai": entry["quai"],
                "Produit": entry["produit"],
                "Début (h)": f"{entry['debut']:.2f}",
                "Fin (h)": f"{entry['fin']:.2f}",
                "Quantité (t)": entry["quantite_chargee"],
                "Axe usine": entry["axe_usine"],
                "Axe port": entry["axe_port"],
                "Source": entry["source"]
            })
        else:
            planning_par_quai[quai].append({
                "Navire": entry["navire"],
                "Produit": entry["produit"],
                "Message": entry.get("message", "Information manquante")
            })

    # Affichage par quai
    for quai, entries in planning_par_quai.items():
        st.markdown(f"### 🧱 {quai}")
        
        # Séparer les cas avec message d'erreur
        entries_valides = [e for e in entries if "Quantité (t)" in e]
        entries_erreurs = [e for e in entries if "Message" in e]

        if entries_valides:
            df = pd.DataFrame(entries_valides)
            st.dataframe(df, use_container_width=True)
        
        for err in entries_erreurs:
            st.warning(f"{err['Navire']} | {err['Produit']} | {err['Message']}")

    # Résumé total
    st.success(f"🚢 Total chargé dans la journée : **{total_charge_journee} tonnes**")
    

elif page == "Stock":
    import pandas as pd
    from pathlib import Path
    from io import BytesIO
    from openpyxl import load_workbook
    import streamlit as st
    
    def extract_jfc_data(file, sheet_name="Feuil1", source_prefix="JFC"):

        wb = load_workbook(filename=file, data_only=True)
        ws = wb[sheet_name]
        data = []
        max_row = 100
        current_source = None
        current_hangar = None
        
        for row in range(2, max_row + 2):
            source_value = ws[f'D{row}'].value
            if source_value and str(source_value).startswith(source_prefix):
                current_source = source_value
            
            if current_source:
                hangar = ws[f'E{row}'].value
                if hangar:
                    current_hangar = hangar
                
                product_type = ws[f'G{row}'].value
                quantity = ws[f'H{row}'].value
                
                if product_type and quantity:
                    normalized_product = normalize_product_name(product_type)
                    data.append({
                        "Source": current_source,
                        "Type de produit": normalized_product,
                        "Quantité": quantity
                    })
        
        return data
    def extract_107_data(file, sheet_name="Feuil1", source_prefix="107"):
        wb = load_workbook(filename=file, data_only=True)
        ws = wb[sheet_name]
        data = []
        max_row = 100
        current_source = None
        current_hangar = None
        for row in range(2, max_row + 2):
            source_value = ws[f'D{row}'].value
            if source_value and str(source_value).startswith(source_prefix):
                current_source = source_value
            
            if current_source:
                hangar = ws[f'E{row}'].value
                if hangar:
                    current_hangar = hangar
                
                product_type = ws[f'G{row}'].value
                quantity = ws[f'H{row}'].value
                
                if product_type and quantity:
                    normalized_product = normalize_product_name(product_type)
                    data.append({
                        "Source": current_source,
                        "Type de produit": normalized_product,
                        "Quantité": quantity
                    })
        
        return data
       

    def read_stock_from_excel(file, sheet_name):
        df = pd.read_excel(file, sheet_name=sheet_name, header=None)
        hangars = df.iloc[19:30, 1].dropna().tolist()
        products = df.iloc[18, 3:27].dropna().tolist()
        
        stock_data = []
        for hangar_index, hangar in enumerate(hangars):
            for product_index, product in enumerate(products):
                quantity = df.iloc[hangar_index + 19, product_index + 3]
                quantity = pd.to_numeric(quantity, errors='coerce')
                if pd.notna(quantity) and quantity > 1000:
                    stock_data.append({
                        "Source": hangar,
                        "Type de produit": product,
                        "Quantité": quantity
                    })
        
        return stock_data
        
    def process_multiple_excels(uploaded_files, sheet_name):
        all_data = []
        for file in uploaded_files:
            all_data.extend(read_stock_from_excel(file, sheet_name))
        return all_data

    def merge_data(jfc_file, jfc_sheet, jfc_prefix, jln_files, stock_sheet, jfc_107_file, jfc_107_sheet):
        all_data = []
        all_data.extend(extract_jfc_data(jfc_file, jfc_sheet, jfc_prefix))
        all_data.extend(process_multiple_excels(jln_files, stock_sheet))
        all_data.extend(extract_107_data(jfc_107_file, jfc_107_sheet, "107"))
        return pd.DataFrame(all_data)
    
    
    def get_background_color(source):
        color_map = {
            "JFC1": "yellow",
            "JFC2": "#1DCD9F",
            "JFC3": "#60B5FF",
            "JFC4": "#C68EFD",
            "JFC5": "#FF9B17",
            "107F": "#379777",
            "107E": "#C68EFD",
            "107D": "#1DCD9F"
        }
        return color_map.get(source, "#B7B7B7")

    def display_dashboard():
        st.title("📦 Suivi des Stocks de Produits Engrais")

    # Initialize session state for uploaded files
        if "jfc_file" not in st.session_state:
            st.session_state["jfc_file"] = None
        if "jln_files" not in st.session_state:
            st.session_state["jln_files"] = None
        if "jfc_107_file" not in st.session_state:
            st.session_state["jfc_107_file"] = None

        # Initialize variables for stock data
        stock_data_jfc = []
        stock_data_jln = []
        stock_data_107 = []

        # File uploaders
        jfc_file = st.file_uploader("📄 Fichier JFC principal", type=["xlsx"])
        if jfc_file:
            st.session_state["jfc_file"] = jfc_file
            stock_data_jfc = extract_jfc_data(jfc_file)

        jln_files = st.file_uploader("📁 Fichiers de stock des JLN", type=["xlsx"], accept_multiple_files=True)
        if jln_files:
            st.session_state["jln_files"] = jln_files
            stock_data_jln = process_multiple_excels(jln_files, "Tableau JLN")
        print(stock_data_jln)       
        jfc_107_file = st.file_uploader("📄 Fichier JFC 107", type=["xlsx"])
        if jfc_107_file:
            st.session_state["jfc_107_file"] = jfc_107_file
            stock_data_107 = extract_107_data(jfc_107_file)

        # Combine stock data
        XSTOCK = stock_data_jfc + stock_data_jln + stock_data_107
        print(XSTOCK)

        if XSTOCK:
            df_stock = pd.DataFrame(XSTOCK)
            output_file = "combined_stock.xlsx"
            df_stock.to_excel(output_file, index=False)
            st.success(f"Stock data saved successfully to {output_file}")
        else:
            st.warning("No stock data available to save.")
        st.session_state["stock_data"] = XSTOCK
        # Check if all files are uploaded
        if st.session_state["jfc_file"] and st.session_state["jln_files"] and st.session_state["jfc_107_file"]:
            # Process and merge data
            df = merge_data(
                st.session_state["jfc_file"], "Feuil1", "JFC",
                st.session_state["jln_files"], "Tableau JLN",
                st.session_state["jfc_107_file"], "Feuil1"
            )

            # Save the processed data in session state
            
            df["Type de produit"] = df["Type de produit"].astype(str)

            # Filtrer par types
            map_df = df[df["Type de produit"].str.startswith("MAP")]
            dap_df = df[df["Type de produit"].str.startswith("DAP")]
            np_df = df[df["Type de produit"].str.startswith("NP")]
            tsp_df = df[df["Type de produit"].str.startswith("TSP")]

            st.write("### 📊 Tableau des stocks de produits par source et quantité")

            columns = st.columns(4)

            def display_product_in_column(df, col, title):
                col.markdown(f"#### {title}")
                products = df["Type de produit"].unique()
                for product in products:
                    product_data = df[df["Type de produit"] == product]
                    total_quantity = product_data["Quantité"].sum()
                    col.markdown(f"##### Produit : {product}")
                    for _, row in product_data.iterrows():
                        source = row["Source"]
                        quantity = row["Quantité"]
                        percentage = (quantity / total_quantity) * 100
                        background_color = get_background_color(source)
                        col.markdown(f'<div style="font-size:14px; padding: 5px; background-color:{background_color}; border-radius: 8px;">'
                                     f'<strong>{source}: </strong> {quantity} t '
                                     f'({percentage:.1f}%)</div>', unsafe_allow_html=True)
                        col.progress(percentage / 100)

            display_product_in_column(map_df, columns[0], "MAP")
            display_product_in_column(dap_df, columns[1], "DAP")
            display_product_in_column(np_df, columns[2], "NP")
            display_product_in_column(tsp_df, columns[3], "TSP")

        else:
            st.warning("Merci d’uploader tous les fichiers nécessaires (JFC principal, JLN, JFC 107).")

    display_dashboard()



elif page == "CTE":
    from io import BytesIO
    import matplotlib.pyplot as plt
    st.title("Suivi des CTE (Contrôle Tirant d'EAU)")

    with st.form("form_cte"):
        # Récupérer la liste des navires disponibles
        navires_disponibles = []
        for quai, info in loading_data.items():
            if info["ship"] and not pd.isna(info["ship"]):
                navires_disponibles.append(info["ship"])

        navire_choisi = st.selectbox("🚢 Choisir un navire", navires_disponibles, key="navire_choisi")

        # Si le navire sélectionné a changé, réinitialiser la sélection de qualité
        if "prev_navire" not in st.session_state:
            st.session_state.prev_navire = navire_choisi
        elif st.session_state.prev_navire != navire_choisi:
            old_key = f"qualite_choisie_{st.session_state.prev_navire}"
            if old_key in st.session_state:
                del st.session_state[old_key]
            st.session_state.prev_navire = navire_choisi

        # Calculer la liste des qualités associées au navire choisi
        qualites = []
        for quai, info in loading_data.items():
            if info["ship"] == navire_choisi:
                qualites = list(info["products"].keys())
                st.session_state.data_navire = info  # stocker les infos du navire
                break

        # Utiliser une clé dynamique pour la sélection de la qualité
        qualite_choisie = st.selectbox("🌾 Choisir la qualité (produit)", qualites, key=f"qualite_choisie_{navire_choisi}")

        type_cte = st.selectbox("📝 Type de CTE", [
            "Fin de chargement",
            "Changement de Qualité",
            "Changement d'Origine",
            "Vérification de tonnage par le Bord",
            "JPH"
        ])
        # ... (le reste du formulaire)

        valeur_cte = st.number_input("⚖️ Valeur mesurée du CTE (tonnes)", min_value=0.0, format="%.2f")

        submit = st.form_submit_button("✅ Valider")

    if submit:
        # Récupérer les données du suivi à partir des informations stockées
        charge_bascule = st.session_state.data_navire["products"][qualite_choisie]["loaded"]
        target = st.session_state.data_navire["products"][qualite_choisie]["target"]

        ecart = valeur_cte - charge_bascule
        reste_a_charger = target - valeur_cte

        # Création de 2 colonnes : colonne gauche pour le graphique, colonne droite pour les résultats
        col1, col2 = st.columns(2)

        with col1:
            # Graphique
            fig, ax = plt.subplots(figsize=(6, 4))
            ax.bar(
                ["CTE", "Bascule", "Demande"],
                [valeur_cte, charge_bascule, target],
                color=["#1f77b4", "#2ca02c", "#7f7f7f"],
                edgecolor="black",
                linewidth=1.2
            )
            ax.set_ylabel("Tonnage (t)", fontsize=12)
            ax.set_title(f"Comparaison CTE vs Bascule - {navire_choisi}", fontsize=14)
            ax.grid(axis="y", linestyle="--", alpha=0.7)
            st.pyplot(fig)
            buf = BytesIO()
            fig.savefig(buf, format="png")
            buf.seek(0)
            img_data = buf.read()

        with col2:
            # Informations des résultats
            st.markdown(f"### Résultats pour le navire **{navire_choisi}** – Qualité **{qualite_choisie}**")
            st.metric("⚖️ CTE mesuré", f"{valeur_cte} t")
            st.metric("📦 Chargé par bascule", f"{charge_bascule} t")
            st.metric("🔀 Écart CTE vs Bascule", f"{ecart:+.2f} t")
            st.metric("📈 Reste à charger", f"{reste_a_charger:.2f} t")
     
        # Convertir le graphique en image


        # Envoi de l'e-mail
        import os
        from email.message import EmailMessage
        import smtplib
        import streamlit as st
        from dotenv import load_dotenv
        import smtplib
        from email.message import EmailMessage

        # Charger les variables d’environnement
        load_dotenv()

        EMAIL_ADDRESS = "elhafianimouad@gmail.com"
        EMAIL_PASSWORD = "txlg orhh icba bbjf"

        # Remplir dynamiquement les données
        email_destinataire = "mouad.elhafiani.etu21@ensem.ac.ma"
        message = EmailMessage()
        message["Subject"] = f"📩 Rapport CTE – {navire_choisi}"
        message["From"] = EMAIL_ADDRESS
        message["To"] = email_destinataire

        message.set_content(f"""
        Rapport CTE
        Navire : {navire_choisi}
        Produit : {qualite_choisie}
        Type de CTE : {type_cte}

        📊 Résumé :
        - Valeur CTE : {valeur_cte} t
        - Chargé par bascule : {charge_bascule} t
        - Écart : {ecart:.2f} t
        - Reste à charger : {reste_a_charger:.2f} t
        """)
        html_content = f"""
        <html>
        <body>
            <p>Navire : {navire_choisi}
          Produit : {qualite_choisie}
        Type de CTE : {type_cte}

        📊 Résumé :
        - Valeur CTE : {valeur_cte} t
        - Chargé par bascule : {charge_bascule} t
        - Écart : {ecart:.2f} t
        - Reste à charger : {reste_a_charger:.2f} t</p>
            <p>Image intégrée dans le contenu :</p>
            <p><img src="cid:graphique"></p>
        </body>
        </html>
        """
        message.add_alternative(html_content, subtype="html")
        message.get_payload()[1].add_related(img_data, maintype="image", subtype="png", cid="<graphique>")

    
        try:
            with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as server:
                server.starttls()
                server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)  # Pour Gmail, assurez-vous d'utiliser un mot de passe d'application si 2FA est activé
                server.send_message(message)
            st.success("✉️ E-mail envoyé avec succès !")
        except Exception as e:
            st.error(f"Erreur lors de l'envoi de l'email : {e}")


        # Enregistrer l'événement CTE dans l'historique (les 10 derniers)
        cte_record = {
            "Navire": navire_choisi,
            "Qualité": qualite_choisie,
            "Type de CTE": type_cte,
            "CTE mesuré (t)": valeur_cte,
            "Bascule (t)": charge_bascule,
            "Écart (t)": round(ecart, 2),
            "Reste à charger (t)": round(reste_a_charger, 2),
            "Horodatage": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        }
        if "cte_history" not in st.session_state:
            st.session_state["cte_history"] = []
        st.session_state["cte_history"].append(cte_record)
        st.session_state["cte_history"] = st.session_state["cte_history"][-10:]

        st.markdown("### Dernières 10 CTE effectuées")
        st.table(pd.DataFrame(st.session_state["cte_history"]))

elif page == "Navires en Rade":
    
    import pandas as pd
    from pathlib import Path
    from io import BytesIO
    from openpyxl import load_workbook
    import matplotlib.pyplot as plt

    def extract_all_ships_data1(file, sheet_name="Feuil1"):
        """Extracts data for all waiting ships in an Excel file, grouped by normalized product type."""
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb[sheet_name]
        data1 = []
        print(f"Max row: {ws.max_row}")

        # Find "I-Navires chargés/Déchargés"
        start_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "I-Navires chargés/Déchargés":
                    start_row = cell.row
                    break
            if start_row:
                break
        if not start_row:
            print("Section 'I-Navires chargés/Déchargés' not found.")
            return data1

        # Now find "III- Navires en attente"
        start_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "III- Navires en attente":
                    start_row = cell.row
                    break
            if start_row:
                break
        if not start_row:
            print("Section 'III- Navires en attente' not found.")
            return data1

        # Locate the "Navires" header in column C
        navire_start_row = None
        for row in ws.iter_rows(min_row=start_row+1, max_row=ws.max_row+1, min_col=3, max_col=3):
            for cell in row:
                if cell.value == "Navires":
                    navire_start_row = cell.row + 2  # data starts two rows below
                    break
            if navire_start_row:
                break
        if not navire_start_row:
            print("Header 'Navires' not found for waiting ships.")
            return data1

        row = navire_start_row  # start processing from here
        current_product = None
        waiting_ship_data = []
        current_ship = None

        while row <= ws.max_row:
            # Read raw product type and normalize it once
            raw_product = ws[f'L{row}'].value
            normalized_product = normalize_product_name(raw_product) if raw_product else None
            print(f"Processing row {row}, normalized product type: {normalized_product}")

            if not raw_product:
                break

            # If a new product type is encountered
            if normalized_product != current_product:
                
                if current_product is not None:
                    data1.append({
                        "Product Type": current_product,
                        "Waiting Ship Data": waiting_ship_data
                    })
                current_product = normalized_product
                waiting_ship_data = []

            # Get other ship-related info
            navire = ws[f'C{row}'].value
            if navire:
                current_ship = navire

            quantity_requested = ws[f'E{row}'].value
            arrival_date = ws[f'H{row}'].value
            destination = ws[f'K{row}'].value

            if current_ship and (quantity_requested or arrival_date or destination):
                normalized_product = normalize_product_name(current_product)
                
                # Use the current normalized product (do not re-normalize)
                
                waiting_ship_data.append({
                    "Navire": current_ship,
                    "Type de produit": normalized_product,
                    "Quantité demandée": quantity_requested if quantity_requested is not None else "N/A",
                    "Date d'arrivée": arrival_date if arrival_date is not None else "N/A",
                    "Destination": destination if destination is not None else "N/A"
                })
            row += 1

        # Append any remaining group
        if current_product:
            # Normalize the last product type
            normalized_product = normalize_product_name(current_product)
            data1.append({
                "Product Type": normalized_product,
                "Waiting Ship Data": waiting_ship_data
            })
        return data1 
    
    def extract_all_ships_data2(file, sheet_name="Feuil1"):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb[sheet_name]
        data2 = []
        start_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "I-Navires chargés/Déchargés":
                    start_row = cell.row
                    break
            if start_row:
                break
        
        if not start_row:
            return data2
        
        navire_start_row = None
        for row in ws.iter_rows(min_row=start_row + 4, max_row=ws.max_row + 1, min_col=3, max_col=3):
            for cell in row:
                if cell.value == "Navires":
                    navire_start_row = cell.row 
                    break
            if navire_start_row:
                break
        
        if not navire_start_row:
            return data2
        
        row = navire_start_row + 1
        current_product = None
        product_data = []
        current_ship = None
        current_quai = None
        while row <= ws.max_row:
            product_type = ws[f'L{row}'].value  
            if not product_type:
                break
            
            if product_type != current_product:
                if current_product:
                    normalized_product = normalize_product_name(current_product)
                    data2.append({
                        "Product Type": normalized_product,
                        "Ship Data": product_data
                    })
                current_product = product_type
                product_data = []
            
            navire = ws[f'C{row}'].value
            if navire:
                current_ship = navire
            
            quai = ws[f'D{row}'].value
            if quai:
                current_quai = quai
            
            quantity_requested = ws[f'E{row}'].value
            tonnage_cumule_7am = ws[f'K{row}'].value or 0
            tonnage_7am = ws[f'J{row}'].value or 0
            rest_a_charger = quantity_requested - tonnage_cumule_7am if quantity_requested else 0
            
            if current_ship and current_quai:
                if quantity_requested or tonnage_cumule_7am or tonnage_7am:
                    normalized_product = normalize_product_name(product_type)
                    product_data.append({
                        "Navire": current_ship,
                        "Quai": current_quai,
                        "Type_Produit": normalized_product,
                        "Reste_A_Char": rest_a_charger
                    })
            row += 1
        
        if current_product:
            normalized_product = normalize_product_name(current_product)
            data2.append({
                "Product Type": normalized_product,
                "Ship Data": product_data
            })
        return data2

            

    # Example usage
    file = st.file_uploader("📄 Upload the Excel file for ship data", type=["xlsx"])
    def save_data_to_excel(data, output_file):
        """
        Sauvegarde les données extraites dans un fichier Excel.
        :param data: Liste des données extraites par extract_all_ships_data.
        :param output_file: Chemin du fichier Excel de sortie.
        """
        # Préparer les données pour l'exportation
        rows = []
        for entry in data:
            
            product_type = entry["Product Type"]
            normalized_product = normalize_product_name(product_type)
            for ship in entry["Waiting Ship Data"]:              
                rows.append({
                    "Type de produit": normalized_product,
                    "Navire": ship["Navire"],
                    "Quantité demandée": ship["Quantité demandée"],
                    "Date d'arrivée": ship["Date d'arrivée"],
                    "Destination": ship["Destination"]
                })

        # Convertir en DataFrame
        df = pd.DataFrame(rows)

        # Sauvegarder dans un fichier Excel
        df.to_excel(output_file, index=False)
        print(f"Les données ont été enregistrées dans le fichier : {output_file}")
    if file is not None:
        data1 = extract_all_ships_data1(file)
        st.session_state['navires_data1'] = data1
        output_file = "waiting_ships_data.xlsx"
        save_data_to_excel(data1, output_file)
        data2 = extract_all_ships_data2(file)
        st.session_state['navires_data'] = data2 
        print(data2)
    else:
        st.warning("Please upload a ship data Excel file.") 
    


    # Sauvegarder les résultats dans un fichier Excel

    
    def charger_donnees(navires_path, stock_path):
        navires = pd.read_excel(navires_path)
        stock = pd.read_excel(stock_path)

        # 🔧 Conversion de la date d'arrivée en datetime
       

        return navires, stock   

    def selectionner_meilleurs_navires(navires, stock):
        """
        Sélectionne les meilleurs navires à charger en fonction des stocks disponibles.
        Chaque navire peut être chargé avec différents types de produits depuis plusieurs sources.
        """
        # Créer une liste pour stocker les résultats
        resultats = []
        for _, navire in navires.iterrows():
            # Convert "Quantité demandée" to a numeric value
            try:
                quantite_demandee = float(navire['Quantité demandée'])
            except (ValueError, TypeError):
                # Skip this row if conversion fails
                continue

            date_arrivee = navire["Date d'arrivée"]
            destination = navire["Destination"]
            navire_nom = navire["Navire"]
            type_produit = navire["Type de produit"]

            stocks_disponibles = stock[stock["Type de produit"] == type_produit].copy()
            stocks_disponibles = stocks_disponibles.sort_values(by="Quantité", ascending=False)

            charge_possible = 0
            sources_utilisees = []

            for _, source in stocks_disponibles.iterrows():
                if charge_possible >= quantite_demandee:
                    break

                quantite_disponible = source["Quantité"]
                source_nom = source["Source"]

                # Calculate the amount from this source
                quantite_a_charger = min(quantite_demandee - charge_possible, quantite_disponible)
                charge_possible += quantite_a_charger
                sources_utilisees.append({
                    "Source": source_nom,
                    "Quantité chargée": quantite_a_charger
                })

            resultats.append({
                "Navire": navire_nom,
                "Type de produit": type_produit,
                "Quantité demandée": quantite_demandee,
                "Charge_possible": charge_possible,
                "Date d'arrivée": date_arrivee,
                "Destination": destination,
                "Sources": sources_utilisees
            })

        # Flatten results into a DataFrame
        resultats_expandes = []
        for resultat in resultats:
            for source in resultat["Sources"]:
                resultats_expandes.append({
                    "Navire": resultat["Navire"],
                    "Type de produit": resultat["Type de produit"],
                    "Quantité demandée": resultat["Quantité demandée"],
                    "Charge_possible": source["Quantité chargée"],
                    "Source": source["Source"],
                    "Date d'arrivée": resultat["Date d'arrivée"],
                    "Destination": resultat["Destination"]
                })
        return pd.DataFrame(resultats_expandes)

    # Exemple d'utilisation
    navires_path = "waiting_ships_data.xlsx"
    stock_path = "combined_stock.xlsx"

    # Charger les données
    navires, stock = charger_donnees(navires_path, stock_path)

    # Sélectionner les meilleurs navires

    meilleurs_navires = selectionner_meilleurs_navires(navires, stock)
    print(selectionner_meilleurs_navires(navires, stock))
    # --------- Affichage Streamlit ---------
    st.title("🚢 Sélection des Meilleurs Navires à Charger")
    st.markdown("Affichage des navires optimaux avec les meilleures combinaisons **Source - Produit - Quantité disponible**.")

    if not meilleurs_navires.empty:
        for navire in meilleurs_navires['Navire'].unique():
            data_navire = meilleurs_navires[meilleurs_navires['Navire'] == navire]

            with st.container(border=True):
                st.markdown(f"## 🛳️ **{navire}**")

                # Destination
                st.markdown(f"**📍 Destination** : {data_navire['Destination'].iloc[0]}")

                # Formatage de la date d'arrivée
                date_arrivee = data_navire["Date d'arrivée"].iloc[0]
                if pd.notnull(date_arrivee):
                    date_formatee = date_arrivee
                else:
                    date_formatee = "Non disponible"
                st.markdown(f"**📅 Date d'arrivée** : {date_formatee}")

                # Détails de chargement par source et produit
                st.markdown("#### 🔄 Détails des chargements :")
                st.table(data_navire.rename(columns={
                    'Type de produit': 'Produit',
                    'Quantité demandée': 'Demandé (t)',
                    'Charge_possible': 'Chargé (t)',
                    'Source': 'Source'
                })[['Produit', 'Demandé (t)', 'Chargé (t)', 'Source']])
    else:
        st.warning("Aucun navire optimal n’a été identifié.")

    
    def flatten_waiting_ships(waiting_data):
        """
        Transforme la liste des données extraites des navires en attente (data1)
        en un DataFrame contenant une ligne par navire.
        """
        rows = []
        for entry in waiting_data:
            product_type = entry.get("Product Type")
            for ship in entry.get("Waiting Ship Data", []):
                rows.append({
                    "Type de produit": product_type,
                    "Navire": ship.get("Navire"),
                    "Quantité demandée": ship.get("Quantité demandée"),
                    "Date d'arrivée": ship.get("Date d'arrivée"),
                    "Destination": ship.get("Destination")
                })
        return pd.DataFrame(rows)
    if 'navires_data1' in st.session_state:
        waiting_ships_df = flatten_waiting_ships(st.session_state['navires_data1'])
        with st.expander("📊 Voir le tableau complet des navires en rade"):
            st.dataframe(waiting_ships_df, use_container_width=True)
    else:
        st.warning("Aucune donnée des navires en rade n'est disponible. Veuillez uploader le fichier Excel.")